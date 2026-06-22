import os
import json
import csv
import threading
import time
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from fastapi import FastAPI, BackgroundTasks, HTTPException
from fastapi.responses import FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
import uvicorn

from screener_logic import run_sepa_screener, sanitize_value

app = FastAPI(title="VN Stock Screener & Scoring API")

# File locations
CACHE_RESULTS_FILE = "screener_results_cache.json"
PREV_RESULTS_FILE = "previous_screener_results.json"
ALERTS_LOG_FILE = "alerts_log.json"

# State variables
screener_state = {
    "is_running": False,
    "progress_pct": 0.0,
    "progress_msg": "Sẵn sàng lọc cổ phiếu",
    "last_update": None
}

state_lock = threading.Lock()

# Update the state of the scanner
def update_state(is_running, pct, msg):
    with state_lock:
        screener_state["is_running"] = is_running
        screener_state["progress_pct"] = pct
        screener_state["progress_msg"] = msg

# Progress callback for the screener_logic
def progress_callback(msg, pct):
    update_state(True, pct, msg)

# Background runner for the screener
def run_screener_background():
    global screener_state
    try:
        update_state(True, 0.0, "Khởi chạy bộ quét...")
        
        # 1. Run the screener
        results = run_sepa_screener(progress_callback=progress_callback)
        
        # 2. Alerts logic (compare with previous session)
        prev_results = {}
        if os.path.exists(PREV_RESULTS_FILE):
            try:
                with open(PREV_RESULTS_FILE, "r", encoding="utf-8") as f:
                    prev_list = json.load(f)
                    prev_results = {x['ticker']: x for x in prev_list}
            except Exception as e:
                print(f"Error loading previous results: {e}")
        
        new_alerts = []
        timestamp_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        for r in results:
            ticker = r['ticker']
            score = r['score']
            
            # Alert 1: Score reaches >= 7 for the first time
            prev_score = prev_results.get(ticker, {}).get('score', 0)
            if score >= 7 and prev_score < 7:
                new_alerts.append({
                    "time": timestamp_str,
                    "ticker": ticker,
                    "score": score,
                    "type": "Tăng Điểm (>7)",
                    "details": f"Điểm số tăng vượt bậc lên {score} điểm (phiên trước: {prev_score} điểm). Xếp hạng: {r['classification']}."
                })
            
            # Alert 2: Breakout with high volume
            if r.get('is_breakout'):
                new_alerts.append({
                    "time": timestamp_str,
                    "ticker": ticker,
                    "score": score,
                    "type": "Breakout Vol lớn",
                    "details": f"Giá bứt phá vượt đỉnh ngắn hạn với khối lượng nổ gấp {r['breakout_vol_ratio']} lần trung bình 20 phiên."
                })
        
        # Load and append alerts
        alerts_log = []
        if os.path.exists(ALERTS_LOG_FILE):
            try:
                with open(ALERTS_LOG_FILE, "r", encoding="utf-8") as f:
                    alerts_log = json.load(f)
            except Exception:
                pass
                
        if new_alerts:
            alerts_log = new_alerts + alerts_log
            alerts_log = alerts_log[:200]  # Limit to 200 logs
            with open(ALERTS_LOG_FILE, "w", encoding="utf-8") as f:
                json.dump(alerts_log, f, ensure_ascii=False, indent=2)
                
        # Save current results as reference for next session
        with open(PREV_RESULTS_FILE, "w", encoding="utf-8") as f:
            json.dump(results, f, ensure_ascii=False, indent=2)
            
        # Save cache
        with open(CACHE_RESULTS_FILE, "w", encoding="utf-8") as f:
            json.dump(results, f, ensure_ascii=False, indent=2)
            
        update_state(False, 1.0, f"Hoàn thành! Đã tính điểm cho {len(results)} cổ phiếu.")
        
    except Exception as e:
        print(f"Error in background scanner: {e}")
        update_state(False, 0.0, f"Lỗi hệ thống: {e}")

# Check and perform market-close auto update
def check_market_close_auto_update():
    while True:
        try:
            now = datetime.now()
            # Run only on weekdays, after 16:00
            is_weekday = now.weekday() < 5
            is_after_market = now.hour >= 16
            
            last_update_date = None
            if os.path.exists(CACHE_RESULTS_FILE):
                try:
                    mtime = os.path.getmtime(CACHE_RESULTS_FILE)
                    last_update_date = datetime.fromtimestamp(mtime).date()
                except Exception:
                    pass
            
            today = now.date()
            if is_weekday and is_after_market and (last_update_date is None or last_update_date < today):
                with state_lock:
                    is_running = screener_state["is_running"]
                if not is_running:
                    print("Auto-triggering daily market close scan...")
                    t = threading.Thread(target=run_screener_background)
                    t.daemon = True
                    t.start()
        except Exception as e:
            print(f"Error in daily update checker: {e}")
            
        # Sleep for 15 minutes before checking again
        time.sleep(900)

# Start auto-update checker thread on startup
@app.on_event("startup")
def startup_event():
    t = threading.Thread(target=check_market_close_auto_update)
    t.daemon = True
    t.start()

# API Endpoints
@app.post("/api/screener/start")
def start_screener():
    with state_lock:
        is_running = screener_state["is_running"]
    if is_running:
        return {"status": "error", "message": "Quá trình quét đang được thực hiện."}
    
    t = threading.Thread(target=run_screener_background)
    t.daemon = True
    t.start()
    return {"status": "success", "message": "Bắt đầu quét cổ phiếu ở chế độ nền."}

@app.get("/api/screener/status")
def get_screener_status():
    with state_lock:
        return screener_state

@app.get("/api/screener/results")
def get_screener_results():
    if not os.path.exists(CACHE_RESULTS_FILE):
        return []
    try:
        with open(CACHE_RESULTS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Không thể đọc file cache: {e}")

@app.get("/api/screener/alerts")
def get_screener_alerts():
    if not os.path.exists(ALERTS_LOG_FILE):
        return []
    try:
        with open(ALERTS_LOG_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Không thể đọc file cảnh báo: {e}")

@app.post("/api/screener/alerts/clear")
def clear_screener_alerts():
    try:
        if os.path.exists(ALERTS_LOG_FILE):
            os.remove(ALERTS_LOG_FILE)
        return {"status": "success", "message": "Đã xóa toàn bộ lịch sử cảnh báo."}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Không thể xóa lịch sử cảnh báo: {e}")


@app.get("/api/screener/export/csv")
def export_csv():
    if not os.path.exists(CACHE_RESULTS_FILE):
        raise HTTPException(status_code=404, detail="Chưa có dữ liệu lọc để xuất.")
    
    try:
        with open(CACHE_RESULTS_FILE, "r", encoding="utf-8") as f:
            results = json.load(f)
            
        export_file = "screener_export.csv"
        with open(export_file, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow([
                "Mã CP", "Tên Công Ty", "Tổng Điểm", "Phân Loại", "Giá Hiện Tại (VND)",
                "Cách đỉnh 52w (%)", "MA20 (VND)", "MA50 (VND)", "MA200 (VND)",
                "RS vs VNINDEX 3T (%)", "Vol TB 20 Phiên", "Số Ngày Phân Phối",
                "Ngành", "Sàn"
            ])
            
            for r in results:
                writer.writerow([
                    r['ticker'], r['name'], r['score'], r['classification'], r['close'] * 1000,
                    r['dist_high'], r['ma20'] * 1000, r['ma50'] * 1000, r['ma200'] * 1000,
                    r['rs_vs_index_3m'], r['volume'], r['distribution_days'],
                    r['industry'], r['exchange']
                ])
                
        return FileResponse(
            path=export_file, 
            filename=f"screener_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            media_type="text/csv"
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Lỗi xuất CSV: {e}")

@app.get("/api/screener/export/excel")
def export_excel():
    if not os.path.exists(CACHE_RESULTS_FILE):
        raise HTTPException(status_code=404, detail="Chưa có dữ liệu lọc để xuất.")
        
    try:
        with open(CACHE_RESULTS_FILE, "r", encoding="utf-8") as f:
            results = json.load(f)
            
        export_file = "screener_export.xlsx"
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bộ lọc chấm điểm"
        ws.views.sheetView[0].showGridLines = True
        
        headers = [
            "Mã CP", "Tên Doanh Nghiệp", "Tổng Điểm", "Phân Loại", "Giá Hiện Tại (VND)",
            "% Cách Đỉnh 52w", "MA20 (VND)", "MA50 (VND)", "MA200 (VND)",
            "RS vs VNINDEX (3T %)", "Vol TB 20 Phiên", "Số Phiên Phân Phối",
            "Ngành", "Sàn"
        ]
        ws.append(headers)
        
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            
        fills = {
            "Leader mạnh": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "Watchlist ưu tiên": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
            "Trung tính": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
            "Loại bỏ": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        }
        
        fonts = {
            "Leader mạnh": Font(name="Calibri", size=11, color="006100", bold=True),
            "Watchlist ưu tiên": Font(name="Calibri", size=11, color="1F4E78", bold=True),
            "Trung tính": Font(name="Calibri", size=11, color="7F6000"),
            "Loại bỏ": Font(name="Calibri", size=11, color="C00000")
        }
        
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        
        for row_idx, r in enumerate(results, 2):
            row_data = [
                r['ticker'], r['name'], r['score'], r['classification'], r['close'] * 1000,
                r['dist_high'], r['ma20'] * 1000, r['ma50'] * 1000, r['ma200'] * 1000,
                r['rs_vs_index_3m'], r['volume'], r['distribution_days'],
                r['industry'], r['exchange']
            ]
            ws.append(row_data)
            
            for col_num in range(1, len(row_data) + 1):
                cell = ws.cell(row=row_idx, column=col_num)
                cell.border = thin_border
                
                if col_num in [5, 7, 8, 9]:
                    cell.number_format = '#,##0đ'
                elif col_num in [6, 10]:
                    cell.number_format = '0.0'
                elif col_num == 11:
                    cell.number_format = '#,##0'
                    
                if col_num in [1, 3, 4, 12, 14]:
                    cell.alignment = Alignment(horizontal="center")
                elif col_num in [5, 6, 7, 8, 9, 10, 11]:
                    cell.alignment = Alignment(horizontal="right")
                    
            cls = r['classification']
            class_cell = ws.cell(row=row_idx, column=4)
            if cls in fills:
                class_cell.fill = fills[cls]
                class_cell.font = fonts[cls]
                
        for col in ws.columns:
            max_len = 0
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or '')
                if cell.number_format and ('đ' in cell.number_format or ',' in cell.number_format) and isinstance(cell.value, (int, float)):
                    val = f"{cell.value:,.0f}đ"
                max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
            
        wb.save(export_file)
        
        return FileResponse(
            path=export_file,
            filename=f"screener_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Lỗi xuất Excel: {e}")

# Serve UI static files
app.mount("/", StaticFiles(directory="static", html=True), name="static")

if __name__ == "__main__":
    uvicorn.run(app, host="127.0.0.1", port=8000)
