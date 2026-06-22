import flet as ft
import pandas as pd
import threading
import os
import csv
import json
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from screener_logic import run_sepa_screener

# --- App Styling Constants ---
BG_COLOR = "#0B0E14"         # Deep blue-black
CARD_BG = "#151B26"          # Slate blue card background
ACCENT_GREEN = "#00E676"     # Neon Green
ACCENT_BLUE = "#00B0FF"      # Bright Blue
TEXT_PRIMARY = "#FFFFFF"     # White
TEXT_SECONDARY = "#8A99AD"   # Light grey-blue
ERROR_COLOR = "#FF1744"      # Bright Red
WARNING_COLOR = "#FFD600"    # Amber Yellow

CACHE_RESULTS_FILE = "screener_results_cache.json"
PREV_RESULTS_FILE = "previous_screener_results.json"
ALERTS_LOG_FILE = "alerts_log.json"

def main(page: ft.Page):
    page.title = "VN Stock Screener & Scoring System"
    page.theme_mode = ft.ThemeMode.DARK
    page.bgcolor = BG_COLOR
    page.padding = 15
    page.horizontal_alignment = ft.CrossAxisAlignment.CENTER
    page.scroll = ft.ScrollMode.ADAPTIVE
    
    # State variables
    screener_results = []
    filtered_results = []
    alerts_log = []
    
    # Load cached results on startup
    if os.path.exists(CACHE_RESULTS_FILE):
        try:
            with open(CACHE_RESULTS_FILE, "r", encoding="utf-8") as f:
                screener_results = json.load(f)
                filtered_results = list(screener_results)
        except Exception:
            pass
            
    # Load cached alerts on startup
    if os.path.exists(ALERTS_LOG_FILE):
        try:
            with open(ALERTS_LOG_FILE, "r", encoding="utf-8") as f:
                alerts_log = json.load(f)
        except Exception:
            pass

    # Unique Industries list for filter
    industries = ["Tất cả các ngành"]
    if screener_results:
        unique_inds = sorted(list(set(r.get('industry', 'Chưa phân loại') for r in screener_results)))
        industries.extend(unique_inds)

    # UI State elements
    industry_dropdown = ft.Dropdown(
        label="Lọc theo Ngành",
        width=250,
        options=[ft.dropdown.Option(ind) for ind in industries],
        value="Tất cả các ngành"
    )
    industry_dropdown.on_change = lambda e: apply_filters()
    
    search_input = ft.TextField(
        label="Tìm mã chứng khoán",
        width=200,
        height=45,
        border_color=TEXT_SECONDARY,
        focused_border_color=ACCENT_BLUE
    )
    search_input.on_change = lambda e: apply_filters()
    
    score_filter_dropdown = ft.Dropdown(
        label="Lọc theo Điểm số",
        width=220,
        options=[
            ft.dropdown.Option("Tất cả điểm số"),
            ft.dropdown.Option("9-10 điểm (Leader mạnh)"),
            ft.dropdown.Option("7-8 điểm (Watchlist ưu tiên)"),
            ft.dropdown.Option("5-6 điểm (Trung tính)"),
            ft.dropdown.Option("Dưới 5 điểm (Loại bỏ)")
        ],
        value="Tất cả điểm số"
    )
    score_filter_dropdown.on_change = lambda e: apply_filters()

    sort_column = "score"
    sort_ascending = False

    # UI Components
    status_text = ft.Text(value="Sẵn sàng lọc cổ phiếu", color=TEXT_SECONDARY, size=14)
    progress_bar = ft.ProgressBar(width=400, value=0, color=ACCENT_GREEN, bgcolor="#263238")
    progress_container = ft.Container(
        content=ft.Column([status_text, progress_bar], horizontal_alignment=ft.CrossAxisAlignment.CENTER),
        visible=False
    )
    
    run_btn = ft.Button(
        content=ft.Text("BẮT ĐẦU QUÉT & CHẤM ĐIỂM", weight=ft.FontWeight.BOLD),
        bgcolor=ACCENT_GREEN,
        color=BG_COLOR,
        height=50,
        icon=ft.Icons.PLAY_ARROW,
        style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=10)),
        on_click=lambda e: start_screening_thread()
    )
    
    export_csv_btn = ft.Button(
        content=ft.Text("XUẤT FILE CSV", weight=ft.FontWeight.BOLD),
        bgcolor=ACCENT_BLUE,
        color=TEXT_PRIMARY,
        height=45,
        icon=ft.Icons.DOWNLOAD,
        on_click=lambda e: export_to_csv()
    )
    
    export_excel_btn = ft.Button(
        content=ft.Text("XUẤT FILE EXCEL", weight=ft.FontWeight.BOLD),
        bgcolor=ACCENT_BLUE,
        color=TEXT_PRIMARY,
        height=45,
        icon=ft.Icons.TABLE_CHART,
        on_click=lambda e: export_to_excel()
    )
    
    export_container = ft.Row(
        [export_csv_btn, export_excel_btn],
        alignment=ft.MainAxisAlignment.CENTER,
        visible=len(screener_results) > 0
    )

    # Result Table Layout
    results_table = ft.DataTable(
        bgcolor=CARD_BG,
        border=ft.Border.all(1, "#263238"),
        border_radius=10,
        heading_row_color="#1F2836",
        horizontal_lines=ft.BorderSide(1, "#263238"),
        columns=[
            ft.DataColumn(ft.Text("Mã CP", weight=ft.FontWeight.BOLD), on_sort=lambda e: sort_data("ticker")),
            ft.DataColumn(ft.Text("Điểm số", weight=ft.FontWeight.BOLD), numeric=True, on_sort=lambda e: sort_data("score")),
            ft.DataColumn(ft.Text("Phân Loại", weight=ft.FontWeight.BOLD), on_sort=lambda e: sort_data("classification")),
            ft.DataColumn(ft.Text("Giá Hiện Tại", weight=ft.FontWeight.BOLD), numeric=True, on_sort=lambda e: sort_data("close")),
            ft.DataColumn(ft.Text("Cách Đỉnh 52w", weight=ft.FontWeight.BOLD), numeric=True, on_sort=lambda e: sort_data("dist_high")),
            ft.DataColumn(ft.Text("MA20", weight=ft.FontWeight.BOLD), numeric=True),
            ft.DataColumn(ft.Text("MA50", weight=ft.FontWeight.BOLD), numeric=True),
            ft.DataColumn(ft.Text("MA200", weight=ft.FontWeight.BOLD), numeric=True),
            ft.DataColumn(ft.Text("RS 3T vs Index", weight=ft.FontWeight.BOLD), numeric=True, on_sort=lambda e: sort_data("rs_vs_index_3m")),
            ft.DataColumn(ft.Text("Vol TB 20 Phiên", weight=ft.FontWeight.BOLD), numeric=True, on_sort=lambda e: sort_data("avg_vol_20")),
            ft.DataColumn(ft.Text("Số Ngày PP", weight=ft.FontWeight.BOLD), numeric=True, on_sort=lambda e: sort_data("distribution_days")),
        ],
        rows=[]
    )
    
    empty_state_text = ft.Text(
        "Chưa có dữ liệu lọc. Vui lòng bấm 'BẮT ĐẦU QUÉT & CHẤM ĐIỂM' ở trên.",
        color=TEXT_SECONDARY,
        size=15,
        text_align=ft.TextAlign.CENTER
    )
    
    table_container = ft.Column(
        controls=[results_table],
        scroll=ft.ScrollMode.ALWAYS,
        visible=len(screener_results) > 0
    )
    
    results_summary_text = ft.Text(
        value=f"Tìm thấy {len(filtered_results)} cổ phiếu" if screener_results else "",
        color=TEXT_PRIMARY,
        size=16,
        weight=ft.FontWeight.BOLD
    )

    # Alerts Table Layout
    alerts_table = ft.DataTable(
        bgcolor=CARD_BG,
        border=ft.Border.all(1, "#263238"),
        border_radius=10,
        heading_row_color="#1F2836",
        columns=[
            ft.DataColumn(ft.Text("Thời Gian", weight=ft.FontWeight.BOLD)),
            ft.DataColumn(ft.Text("Mã CP", weight=ft.FontWeight.BOLD)),
            ft.DataColumn(ft.Text("Điểm số", weight=ft.FontWeight.BOLD)),
            ft.DataColumn(ft.Text("Loại Cảnh Báo", weight=ft.FontWeight.BOLD)),
            ft.DataColumn(ft.Text("Chi Tiết Tín Hiệu", weight=ft.FontWeight.BOLD)),
        ],
        rows=[]
    )
    
    clear_alerts_btn = ft.Button(
        content=ft.Text("XÓA LỊCH SỬ CẢNH BÁO", weight=ft.FontWeight.BOLD),
        bgcolor=ERROR_COLOR,
        color=TEXT_PRIMARY,
        height=40,
        icon=ft.Icons.DELETE_SWEEP,
        on_click=lambda e: clear_alerts_log()
    )
    
    alerts_empty_text = ft.Text("Chưa có cảnh báo nào xuất hiện.", color=TEXT_SECONDARY, size=15)
    
    alerts_table_container = ft.Column([alerts_table], scroll=ft.ScrollMode.ALWAYS, visible=False)
    
    alerts_container = ft.Column(
        controls=[
            ft.Row([ft.Text("Danh sách Cảnh báo Mới", size=16, weight=ft.FontWeight.BOLD), clear_alerts_btn], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
            ft.Divider(color="#263238"),
            alerts_empty_text,
            alerts_table_container
        ],
        spacing=10
    )

    # Notification Banner
    def show_snackbar(text, color=ACCENT_BLUE):
        page.snack_bar = ft.SnackBar(
            content=ft.Text(text, color=TEXT_PRIMARY, weight=ft.FontWeight.BOLD),
            bgcolor=color,
            duration=4000
        )
        page.snack_bar.open = True
        page.update()

    def update_progress(msg, value):
        pct = int(value * 100)
        status_text.value = f"[{pct}%] {msg}"
        progress_bar.value = value
        page.update()

    def start_screening_thread():
        table_container.visible = False
        empty_state_text.visible = False
        export_container.visible = False
        results_summary_text.value = ""
        
        progress_bar.visible = True
        status_text.value = "Bắt đầu khởi chạy bộ quét..."
        status_text.color = TEXT_SECONDARY
        progress_container.visible = True
        run_btn.disabled = True
        page.update()
        
        t = threading.Thread(target=run_screener_job)
        t.start()

    def run_screener_job():
        nonlocal screener_results
        try:
            # 1. Chạy bộ lọc
            results = run_sepa_screener(progress_callback=update_progress)
            
            # 2. Xử lý logic cảnh báo bằng cách so sánh kết quả cũ
            prev_results = {}
            if os.path.exists(PREV_RESULTS_FILE):
                try:
                    with open(PREV_RESULTS_FILE, "r", encoding="utf-8") as f:
                        prev_list = json.load(f)
                        prev_results = {x['ticker']: x for x in prev_list}
                except Exception:
                    pass
            
            new_alerts = []
            timestamp_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            for r in results:
                ticker = r['ticker']
                score = r['score']
                
                # Cảnh báo 1: Lần đầu đạt >= 7 điểm
                prev_score = prev_results.get(ticker, {}).get('score', 0)
                if score >= 7 and prev_score < 7:
                    new_alerts.append({
                        "time": timestamp_str,
                        "ticker": ticker,
                        "score": score,
                        "type": "Tăng Điểm (>7)",
                        "details": f"Điểm số tăng vượt bậc lên {score} điểm (phiên trước: {prev_score} điểm). Xếp hạng: {r['classification']}."
                    })
                
                # Cảnh báo 2: Breakout với volume lớn
                if r['is_breakout']:
                    new_alerts.append({
                        "time": timestamp_str,
                        "ticker": ticker,
                        "score": score,
                        "type": "Breakout Vol lớn",
                        "details": f"Giá bứt phá vượt đỉnh ngắn hạn với khối lượng nổ gấp {r['breakout_vol_ratio']} lần trung bình 20 phiên."
                    })
            
            # Cập nhật lịch sử cảnh báo
            if new_alerts:
                nonlocal alerts_log
                alerts_log = new_alerts + alerts_log # Thêm vào đầu danh sách
                # Giới hạn 200 cảnh báo gần nhất
                alerts_log = alerts_log[:200]
                
                with open(ALERTS_LOG_FILE, "w", encoding="utf-8") as f:
                    json.dump(alerts_log, f, ensure_ascii=False, indent=2)
                
                show_snackbar(f"Phát hiện {len(new_alerts)} cảnh báo giao dịch mới!", WARNING_COLOR)
            
            # Lưu kết quả hiện tại làm kết quả tham chiếu cho phiên sau
            with open(PREV_RESULTS_FILE, "w", encoding="utf-8") as f:
                json.dump(results, f, ensure_ascii=False, indent=2)
                
            # Lưu kết quả cache hiện tại
            with open(CACHE_RESULTS_FILE, "w", encoding="utf-8") as f:
                json.dump(results, f, ensure_ascii=False, indent=2)
                
            # Cập nhật state chính
            screener_results = results
            
            # Cập nhật bộ lọc Ngành
            unique_inds = sorted(list(set(x.get('industry', 'Chưa phân loại') for x in results)))
            industry_dropdown.options = [ft.dropdown.Option("Tất cả các ngành")] + [ft.dropdown.Option(ind) for ind in unique_inds]
            
            # Render kết quả và cảnh báo lên UI
            apply_filters()
            render_alerts_table()
            
            progress_container.visible = False
            page.update()
            
        except Exception as e:
            progress_bar.visible = False
            status_text.value = f"Lỗi hệ thống: {e}"
            status_text.color = ERROR_COLOR
            show_snackbar(f"Đã xảy ra lỗi trong quá trình quét: {e}", ERROR_COLOR)
            page.update()
        finally:
            run_btn.disabled = False
            page.update()

    def apply_filters():
        nonlocal screener_results, filtered_results
        if not screener_results:
            table_container.visible = False
            empty_state_text.visible = True
            export_container.visible = False
            page.update()
            return

        selected_ind = industry_dropdown.value
        search_query = search_input.value.strip().upper()
        selected_score_class = score_filter_dropdown.value
        
        temp_list = []
        for r in screener_results:
            # Lọc theo Tìm kiếm
            if search_query and search_query not in r['ticker'].upper():
                continue
                
            # Lọc theo Ngành
            if selected_ind != "Tất cả các ngành" and r.get('industry') != selected_ind:
                continue
                
            # Lọc theo phân loại điểm
            if selected_score_class == "9-10 điểm (Leader mạnh)" and r['score'] < 9:
                continue
            elif selected_score_class == "7-8 điểm (Watchlist ưu tiên)" and (r['score'] < 7 or r['score'] > 8):
                continue
            elif selected_score_class == "5-6 điểm (Trung tính)" and (r['score'] < 5 or r['score'] > 6):
                continue
            elif selected_score_class == "Dưới 5 điểm (Loại bỏ)" and r['score'] >= 5:
                continue
                
            temp_list.append(r)
            
        # Sắp xếp dữ liệu theo cột đã chọn
        def sort_key(x):
            val = x.get(sort_column)
            if val is None:
                return "" if isinstance(sort_column, str) else 0
            return val
            
        temp_list.sort(key=sort_key, reverse=not sort_ascending)
        
        filtered_results = temp_list
        render_table_rows()

    def sort_data(column_name):
        nonlocal sort_column, sort_ascending
        if sort_column == column_name:
            sort_ascending = not sort_ascending
        else:
            sort_column = column_name
            sort_ascending = False
            
        apply_filters()

    def render_table_rows():
        nonlocal filtered_results
        rows = []
        for r in filtered_results:
            # Định dạng khối lượng (M hoặc k)
            vol_formatted = f"{r['volume']/1000000:.2f}M" if r['volume'] >= 1000000 else f"{r['volume']/1000:.0f}k"
            
            # Chọn màu cho phân loại điểm số
            score = r['score']
            if score >= 9:
                class_color = ACCENT_GREEN
            elif score >= 7:
                class_color = ACCENT_BLUE
            elif score >= 5:
                class_color = WARNING_COLOR
            else:
                class_color = TEXT_SECONDARY
                
            rows.append(
                ft.DataRow(
                    cells=[
                        ft.DataCell(ft.Text(r['ticker'], color=class_color, weight=ft.FontWeight.BOLD)),
                        ft.DataCell(ft.Text(f"{r['score']}/10", color=class_color, weight=ft.FontWeight.BOLD)),
                        ft.DataCell(ft.Text(r['classification'], color=class_color)),
                        ft.DataCell(ft.Text(f"{r['close']*1000:,.0f}đ", color=TEXT_PRIMARY)),
                        ft.DataCell(ft.Text(f"{r['dist_high']}%", color=TEXT_PRIMARY)),
                        ft.DataCell(ft.Text(f"{r['ma20']*1000:,.0f}đ", color=TEXT_SECONDARY)),
                        ft.DataCell(ft.Text(f"{r['ma50']*1000:,.0f}đ", color=TEXT_SECONDARY)),
                        ft.DataCell(ft.Text(f"{r['ma200']*1000:,.0f}đ", color=TEXT_SECONDARY)),
                        ft.DataCell(ft.Text(f"{r['rs_vs_index_3m']}%", color=ACCENT_BLUE if r['rs_vs_index_3m'] > 0 else ERROR_COLOR)),
                        ft.DataCell(ft.Text(vol_formatted, color=TEXT_PRIMARY)),
                        ft.DataCell(ft.Text(f"{r['distribution_days']}", color=ERROR_COLOR if r['distribution_days'] > 4 else ACCENT_GREEN)),
                    ]
                )
            )
            
        results_table.rows = rows
        results_summary_text.value = f"Hiển thị {len(filtered_results)} cổ phiếu đạt điều kiện bộ lọc"
        
        if len(filtered_results) > 0:
            table_container.visible = True
            empty_state_text.visible = False
            export_container.visible = True
        else:
            table_container.visible = False
            empty_state_text.visible = True
            empty_state_text.value = "Không tìm thấy cổ phiếu nào khớp cấu hình lọc hiện tại."
            export_container.visible = False
            
        page.update()

    def render_alerts_table():
        nonlocal alerts_log
        if not alerts_log:
            alerts_empty_text.visible = True
            alerts_table_container.visible = False
            page.update()
            return
            
        rows = []
        for a in alerts_log:
            type_color = WARNING_COLOR if a['type'] == "Breakout Vol lớn" else ACCENT_BLUE
            rows.append(
                ft.DataRow(
                    cells=[
                        ft.DataCell(ft.Text(a['time'], color=TEXT_SECONDARY)),
                        ft.DataCell(ft.Text(a['ticker'], color=ACCENT_GREEN, weight=ft.FontWeight.BOLD)),
                        ft.DataCell(ft.Text(f"{a['score']}/10", color=TEXT_PRIMARY, weight=ft.FontWeight.BOLD)),
                        ft.DataCell(ft.Text(a['type'], color=type_color, weight=ft.FontWeight.BOLD)),
                        ft.DataCell(ft.Text(a['details'], color=TEXT_PRIMARY)),
                    ]
                )
            )
            
        alerts_table.rows = rows
        alerts_empty_text.visible = False
        alerts_table_container.visible = True
        page.update()

    def clear_alerts_log():
        nonlocal alerts_log
        alerts_log = []
        if os.path.exists(ALERTS_LOG_FILE):
            try:
                os.remove(ALERTS_LOG_FILE)
            except Exception:
                pass
        render_alerts_table()
        show_snackbar("Đã xóa toàn bộ lịch sử cảnh báo.", ACCENT_GREEN)

    def export_to_csv():
        nonlocal filtered_results
        if not filtered_results:
            show_snackbar("Không có dữ liệu để xuất file CSV.", ERROR_COLOR)
            return
            
        try:
            filename = f"screener_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            filepath = os.path.join(os.getcwd(), filename)
            
            with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                writer.writerow([
                    "Mã CP", "Tên Công Ty", "Tổng Điểm", "Phân Loại", "Giá Hiện Tại (VND)",
                    "Cách đỉnh 52w (%)", "MA20 (VND)", "MA50 (VND)", "MA200 (VND)",
                    "RS vs VNINDEX 3T (%)", "Vol TB 20 Phiên", "Số Ngày Phân Phối",
                    "Ngành", "Sàn"
                ])
                
                for r in filtered_results:
                    writer.writerow([
                        r['ticker'], r['name'], r['score'], r['classification'], r['close'] * 1000,
                        r['dist_high'], r['ma20'] * 1000, r['ma50'] * 1000, r['ma200'] * 1000,
                        r['rs_vs_index_3m'], r['volume'], r['distribution_days'],
                        r['industry'], r['exchange']
                    ])
                    
            show_snackbar(f"Xuất file CSV thành công: {filename}", ACCENT_GREEN)
        except Exception as e:
            show_snackbar(f"Không thể xuất file CSV: {e}", ERROR_COLOR)

    def export_to_excel():
        nonlocal filtered_results
        if not filtered_results:
            show_snackbar("Không có dữ liệu để xuất file Excel.", ERROR_COLOR)
            return
            
        try:
            filename = f"screener_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(os.getcwd(), filename)
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Bộ lọc chấm điểm"
            ws.views.sheetView[0].showGridLines = True
            
            headers = [
                "Mã CP", "Tên Doanh Nghiệp", "Tổng Điểm", "Phân Loại", "Giá Hiện Tại (VND)",
                "% Cách Đỉnh 52w", "MA20 (VND)", "MA50 (VND)", "MA200 (VND)",
                "RS vs VNINDEX (3T %)", "Vol TB 20 Phiên", "Số Phiên Phân Phối",
                "Ngành", "Sàn"
            ]
            ws.append(headers)
            
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
            header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                
            fills = {
                "Leader mạnh": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
                "Watchlist ưu tiên": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
                "Trung tính": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
                "Loại bỏ": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
            }
            
            fonts = {
                "Leader mạnh": Font(name="Calibri", size=11, color="006100", bold=True),
                "Watchlist ưu tiên": Font(name="Calibri", size=11, color="1F4E78", bold=True),
                "Trung tính": Font(name="Calibri", size=11, color="7F6000"),
                "Loại bỏ": Font(name="Calibri", size=11, color="C00000")
            }
            
            thin_border = Border(
                left=Side(style='thin', color='D9D9D9'),
                right=Side(style='thin', color='D9D9D9'),
                top=Side(style='thin', color='D9D9D9'),
                bottom=Side(style='thin', color='D9D9D9')
            )
            
            for row_idx, r in enumerate(filtered_results, 2):
                row_data = [
                    r['ticker'], r['name'], r['score'], r['classification'], r['close'] * 1000,
                    r['dist_high'], r['ma20'] * 1000, r['ma50'] * 1000, r['ma200'] * 1000,
                    r['rs_vs_index_3m'], r['volume'], r['distribution_days'],
                    r['industry'], r['exchange']
                ]
                ws.append(row_data)
                
                for col_num in range(1, len(row_data) + 1):
                    cell = ws.cell(row=row_idx, column=col_num)
                    cell.border = thin_border
                    
                    if col_num in [5, 7, 8, 9]:
                        cell.number_format = '#,##0đ'
                    elif col_num in [6, 10]:
                        cell.number_format = '0.0'
                    elif col_num == 11:
                        cell.number_format = '#,##0'
                        
                    if col_num in [1, 3, 4, 12, 14]:
                        cell.alignment = Alignment(horizontal="center")
                    elif col_num in [5, 6, 7, 8, 9, 10, 11]:
                        cell.alignment = Alignment(horizontal="right")
                        
                # Style class cell
                cls = r['classification']
                class_cell = ws.cell(row=row_idx, column=4)
                if cls in fills:
                    class_cell.fill = fills[cls]
                    class_cell.font = fonts[cls]
                    
            for col in ws.columns:
                max_len = 0
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                for cell in col:
                    val = str(cell.value or '')
                    if cell.number_format and ('đ' in cell.number_format or ',' in cell.number_format) and isinstance(cell.value, (int, float)):
                        val = f"{cell.value:,.0f}đ"
                    max_len = max(max_len, len(val))
                ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
                
            wb.save(filepath)
            show_snackbar(f"Xuất file Excel thành công: {filename}", ACCENT_GREEN)
        except Exception as e:
            show_snackbar(f"Không thể xuất file Excel: {e}", ERROR_COLOR)

    # --- Daily Market Close Auto Update Checker ---
    def check_market_close_auto_update():
        now = datetime.now()
        is_weekday = now.weekday() < 5
        is_after_market = now.hour >= 16
        
        last_update_date = None
        if os.path.exists(CACHE_RESULTS_FILE):
            try:
                mtime = os.path.getmtime(CACHE_RESULTS_FILE)
                last_update_date = datetime.fromtimestamp(mtime).date()
            except Exception:
                pass
                
        today = now.date()
        # Kích hoạt quét tự động nếu là ngày trong tuần, sau 16h và chưa cập nhật phiên hôm nay
        if is_weekday and is_after_market and (last_update_date is None or last_update_date < today):
            show_snackbar("Đã qua 16:00. Tự động quét cập nhật dữ liệu thị trường đóng cửa...", ACCENT_BLUE)
            start_screening_thread()

    # --- Tabs Views Layout ---
    
    # TAB 1: BỘ LỌC CỔ PHIẾU
    screener_tab = ft.Column(
        controls=[
            ft.Container(
                content=ft.Row([run_btn, export_container], alignment=ft.MainAxisAlignment.CENTER),
                margin=ft.Margin.symmetric(vertical=10)
            ),
            progress_container,
            ft.Row([
                results_summary_text,
                ft.Row([
                    search_input,
                    industry_dropdown,
                    score_filter_dropdown
                ], alignment=ft.MainAxisAlignment.END, spacing=10)
            ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
            ft.Divider(color="#263238"),
            empty_state_text,
            table_container
        ],
        spacing=10
    )

    # TAB 2: CẤU HÌNH TIÊU CHÍ (DOCUMENTATION)
    criteria_guide_tab = ft.Column(
        controls=[
            ft.Container(
                content=ft.Column([
                    ft.Text("Hệ thống chấm điểm 0–10 điểm - Quy tắc kỹ thuật và cơ bản", size=18, color=ACCENT_GREEN, weight=ft.FontWeight.BOLD),
                    ft.Divider(color="#263238"),
                    ft.Text("Mỗi tiêu chí đạt được dưới đây cộng 1 điểm cho cổ phiếu:", size=14, color=TEXT_PRIMARY),
                    
                    ft.Text("• [1] Giá đóng cửa hiện tại > MA20 (Xu hướng ngắn hạn hướng lên).", size=13, color=TEXT_SECONDARY),
                    ft.Text("• [2] MA20 > MA50 (Đường MA ngắn hạn nằm trên đường MA trung hạn).", size=13, color=TEXT_SECONDARY),
                    ft.Text("• [3] MA50 > MA200 (Đường MA trung hạn nằm trên đường MA dài hạn - Giao cắt vàng).", size=13, color=TEXT_SECONDARY),
                    ft.Text("• [4] MA200 có độ dốc dương trong ít nhất 20 phiên gần nhất (Đảm bảo xu hướng tăng dài hạn vững chắc).", size=13, color=TEXT_SECONDARY),
                    ft.Text("• [5] Giá hiện tại cách đỉnh 52 tuần không quá 15% (Cổ phiếu nằm gần vùng đỉnh lịch sử/52 tuần thể hiện sức mạnh vượt trội).", size=13, color=TEXT_SECONDARY),
                    ft.Text("• [6] Relative Strength (RS) cao hơn VNINDEX trong 3 tháng gần nhất (Mạnh hơn thị trường chung).", size=13, color=TEXT_SECONDARY),
                    ft.Text("• [7] Khối lượng giao dịch TB 20 phiên cao hơn TB 60 phiên (Dòng tiền gia tăng tích cực).", size=13, color=TEXT_SECONDARY),
                    ft.Text("• [8] Có nền giá tích lũy từ 6–12 tuần với biên độ dao động nhỏ hơn 15% (Quá trình siết chặt cạn cung).", size=13, color=TEXT_SECONDARY),
                    ft.Text("• [9] Số phiên phân phối trong 25 phiên gần nhất không vượt quá 4 phiên (Lực bán tháo yếu hoặc được hấp thụ hết).", size=13, color=TEXT_SECONDARY),
                    ft.Text("• [10] Doanh nghiệp có tăng trưởng EPS hoặc lợi nhuận sau thuế YoY dương trong cả 4 quý gần nhất (Nền tảng cơ bản khỏe mạnh).", size=13, color=TEXT_SECONDARY),
                ]),
                bgcolor=CARD_BG,
                padding=20,
                border_radius=10,
                border=ft.Border.all(1, "#263238")
            ),
            ft.Container(
                content=ft.Column([
                    ft.Text("Quy tắc Phân Loại Cổ Phiếu", size=18, color=ACCENT_BLUE, weight=ft.FontWeight.BOLD),
                    ft.Divider(color="#263238"),
                    ft.Row([
                        ft.Icon(ft.Icons.STAR, color=ACCENT_GREEN),
                        ft.Text("9–10 điểm: Leader mạnh", color=ACCENT_GREEN, weight=ft.FontWeight.BOLD)
                    ]),
                    ft.Row([
                        ft.Icon(ft.Icons.REMOVE_RED_EYE, color=ACCENT_BLUE),
                        ft.Text("7–8 điểm: Watchlist ưu tiên", color=ACCENT_BLUE, weight=ft.FontWeight.BOLD)
                    ]),
                    ft.Row([
                        ft.Icon(ft.Icons.HORIZONTAL_RULE, color=WARNING_COLOR),
                        ft.Text("5–6 điểm: Trung tính", color=WARNING_COLOR)
                    ]),
                    ft.Row([
                        ft.Icon(ft.Icons.DELETE_OUTLINE, color=TEXT_SECONDARY),
                        ft.Text("Dưới 5 điểm: Loại bỏ", color=TEXT_SECONDARY)
                    ]),
                ]),
                bgcolor=CARD_BG,
                padding=20,
                border_radius=10,
                border=ft.Border.all(1, "#263238")
            )
        ],
        spacing=15
    )

    # Render data on load
    if screener_results:
        apply_filters()
    render_alerts_table()
    
    # Kích hoạt bộ kiểm tra quét tự động sau đóng cửa
    check_market_close_auto_update()

    # Main Tabs Control
    tabs = ft.Tabs(
        length=3,
        selected_index=0,
        animation_duration=300,
        content=ft.Column(
            controls=[
                ft.TabBar(
                    tabs=[
                        ft.Tab(label="BỘ LỌC CỔ PHIẾU", icon=ft.Icons.FILTER_ALT),
                        ft.Tab(label="CẢNH BÁO MỚI", icon=ft.Icons.NOTIFICATIONS_ACTIVE),
                        ft.Tab(label="TIÊU CHÍ & PHÂN LOẠI", icon=ft.Icons.MENU_BOOK)
                    ]
                ),
                ft.TabBarView(
                    controls=[
                        screener_tab,
                        alerts_container,
                        criteria_guide_tab
                    ],
                    expand=True
                )
            ]
        ),
        expand=1
    )
    
    # Header bar
    header = ft.Container(
        content=ft.Row([
            ft.Icon(ft.Icons.MONETIZATION_ON, color=ACCENT_GREEN, size=32),
            ft.Text("VN SENSE - STOCK SCREENING & SCORING", size=20, weight=ft.FontWeight.BOLD, color=TEXT_PRIMARY),
        ], alignment=ft.MainAxisAlignment.CENTER),
        padding=10,
        bgcolor=CARD_BG,
        border_radius=8,
        margin=ft.Margin.only(bottom=10)
    )

    page.add(header, tabs)

if __name__ == "__main__":
    ft.run(main)
